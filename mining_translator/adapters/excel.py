"""Excel adapter for .xlsx, .xlsm and .xls files."""

import os
from .base import BaseAdapter, TextBlock


def _detect_format(filepath: str) -> str:
    """Detect Excel format from file header. Returns 'xlsx' or 'xls'."""
    with open(filepath, "rb") as f:
        header = f.read(4)
    if header[:2] == b"PK":
        return "xlsx"
    elif header[:4] == b"\xd0\xcf\x11\xe0":
        return "xls"
    return "xlsx"  # default


class ExcelAdapter(BaseAdapter):
    supported_extensions = [".xlsx", ".xlsm", ".xls"]

    def extract_texts(self, filepath: str) -> list[TextBlock]:
        fmt = _detect_format(filepath)
        if fmt == "xls":
            return self._extract_xls(filepath)
        return self._extract_xlsx(filepath)

    def write_translated(self, filepath: str, blocks: list[TextBlock], output_path: str):
        fmt = _detect_format(filepath)
        # Always output .xlsx format (xlrd can't write .xls)
        self._write_xlsx(filepath, blocks, output_path, fmt == "xls")

    def _extract_xls(self, filepath: str) -> list[TextBlock]:
        import xlrd

        wb = xlrd.open_workbook(filepath)
        blocks = []

        for si, sheet in enumerate(wb.sheets()):
            sheet_name = sheet.name
            for ri in range(sheet.nrows):
                for ci in range(sheet.ncols):
                    cell = sheet.cell(ri, ci)
                    if cell.ctype == 1 and cell.value:  # text type
                        text = str(cell.value).strip()
                        if text:
                            coord = f"{xlrd.colname(ci)}{ri + 1}"
                            blocks.append(TextBlock(
                                id=f"{sheet_name}|{coord}",
                                text=text,
                                context=f"Sheet '{sheet_name}', cell {coord}",
                                meta={"sheet": sheet_name, "coordinate": coord},
                            ))

        return blocks

    def _extract_xlsx(self, filepath: str) -> list[TextBlock]:
        from openpyxl import load_workbook

        wb = load_workbook(filepath)
        blocks = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.strip():
                        blocks.append(TextBlock(
                            id=f"{sheet_name}|{cell.coordinate}",
                            text=cell.value,
                            context=f"Sheet '{sheet_name}', cell {cell.coordinate}",
                            meta={"sheet": sheet_name, "coordinate": cell.coordinate},
                        ))

        wb.close()
        return blocks

    def _write_xlsx(self, filepath: str, blocks: list[TextBlock], output_path: str,
                    from_xls: bool = False):
        from openpyxl import Workbook, load_workbook

        if from_xls:
            # Create new workbook from scratch (xlrd can't write)
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            # Group blocks by sheet
            sheets: dict[str, list[TextBlock]] = {}
            for b in blocks:
                sn = b.meta.get("sheet", "Sheet1")
                sheets.setdefault(sn, []).append(b)

            for si, (sn, sblocks) in enumerate(sheets.items()):
                if si == 0:
                    ws = wb.active
                    ws.title = sn
                else:
                    ws = wb.create_sheet(title=sn)

                for b in sblocks:
                    ws[b.meta["coordinate"]] = b.translated

            wb.save(output_path)
        else:
            # Preserve original formatting for .xlsx files
            wb = load_workbook(filepath)
            lookup = {b.id: b.translated for b in blocks}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        block_id = f"{sheet_name}|{cell.coordinate}"
                        if block_id in lookup:
                            cell.value = lookup[block_id]

            wb.save(output_path)
